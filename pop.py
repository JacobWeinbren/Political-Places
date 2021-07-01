from openpyxl import load_workbook
from pyexcel.cookbook import merge_all_to_a_book
import csv, os

import openpyxl
from pyexcel.cookbook import merge_csv_to_a_book

import openpyxl

data = {}

def readFile(book, sheet, id_col, data_col, year):
	print(year)
	wb = load_workbook(filename = book, data_only = 'True')
	sheet_ranges = wb[sheet]

	iterinput = iter(sheet_ranges.rows)
	next(iterinput)

	for value, row in enumerate(iterinput):
		val = row[data_col].value 
		name  = row[id_col].value
		
		if not name in data:
			data[name] = {}

		if not isinstance(val, int):
			val = int(val.replace(',', ''))

		data[name][year] = val

readFile('input/uk_pop/2009.xlsx', 'Mid-2009', 0, 3, '2009') 
readFile('input/uk_pop/2019.xlsx', 'Mid-2019 Persons', 0, 6, '2019') 

with open('output/pop.csv', 'w') as file:
	writer = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
	writer.writerow(['LSOA', 'Population Change %'])

	for lsoa in data.keys():
		item = data[lsoa]

		change = round(((item['2019'] - item['2009']) / item['2009']) * 100, 2)
		writer.writerow([lsoa, change])


print("Saving as xlsx")

#Writes to an xlsx
merge_csv_to_a_book(['output/pop.csv'], outfilename='output/pop.xlsx')

#Renames sheetname
ss=openpyxl.load_workbook("output/pop.xlsx")
ss_sheet = ss['pop.csv']
ss_sheet.title = 'pop'
ss.save("output/pop.xlsx") 

#Deletes csv
os.remove("output/pop.csv")